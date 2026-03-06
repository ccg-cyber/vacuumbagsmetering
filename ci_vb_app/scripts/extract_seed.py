#!/usr/bin/env python3
"""
scripts/extract_seed.py
Extracts Sheet2 (materials) and Sheet4 (costs) from the Excel workbook
and writes assets/materials.json and assets/costs.json.

Usage:
  python3 scripts/extract_seed.py path/to/Metering_support_2_1.xlsx

Output files are placed in assets/ relative to the script's parent directory.
"""
import json
import sys
import os
from openpyxl import load_workbook

def extract(xlsx_path: str):
    wb = load_workbook(xlsx_path, data_only=True)
    out_dir = os.path.join(os.path.dirname(__file__), '..', 'assets')
    os.makedirs(out_dir, exist_ok=True)

    # ----- Sheet2 → materials.json -----
    ws2 = wb['Sheet2']
    materials = []
    for row in ws2.iter_rows(min_row=3, max_row=ws2.max_row, values_only=True):
        if row[0] is None:
            continue
        materials.append({
            "name": str(row[0]).strip(),
            "col_b": _f(row[1]),
            "g_m2": _f(row[2]),
            "thickness_um": _f(row[3]),
            "col_e": _f(row[4]),
            "col_f": _f(row[5]),
            "weight_factor": _f(row[6]),
        })

    mat_path = os.path.join(out_dir, 'materials.json')
    with open(mat_path, 'w', encoding='utf-8') as f:
        json.dump(materials, f, indent=2)
    print(f"✓ {len(materials)} materials → {mat_path}")

    # ----- Sheet4 → costs.json -----
    ws4 = wb['Sheet4']
    costs = []
    seen = set()
    for row in ws4.iter_rows(min_row=3, max_row=ws4.max_row, values_only=True):
        if row[0] is None or row[1] is None:
            continue
        name = str(row[0]).strip()
        if name in seen:
            continue
        seen.add(name)
        try:
            costs.append({"name": name, "cost": float(row[1])})
        except (TypeError, ValueError):
            pass

    cost_path = os.path.join(out_dir, 'costs.json')
    with open(cost_path, 'w', encoding='utf-8') as f:
        json.dump(costs, f, indent=2)
    print(f"✓ {len(costs)} costs → {cost_path}")

def _f(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python3 extract_seed.py <path_to_xlsx>")
        sys.exit(1)
    extract(sys.argv[1])
